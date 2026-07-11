"""تست‌های خواندن استریمی اکسل: گاردهای محدوده‌ی بادکرده، سقف ردیف و پاریته‌ی نگاشت."""

from __future__ import annotations

import io
import sys
import time
from collections import namedtuple
from pathlib import Path

import pandas as pd

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "src"))

from mktcore.connectors.excel_csv import ExcelCsvConnector  # noqa: E402
from mktcore.connectors.excel_stream import (  # noqa: E402
    EMPTY_RUN_LIMIT,
    StreamResult,
    _collect,
    _densify,
    _xlsb_value_iter,
    finalize_frame,
)
from mktcore.ingest.mapper import SchemaMapper  # noqa: E402
from mktcore.synthetic import generate_synthetic_sales  # noqa: E402


def _xlsx_bytes(rows: list[list], *, header: list[str],
                inflate_to_row: int | None = None) -> bytes:
    """ساخت xlsx در حافظه؛ inflate_to_row با یک سلول خالی dimension را باد می‌کند."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    if inflate_to_row:
        # لمس یک سلول دور، بدون مقدار — دقیقاً الگوی فایل‌های با فرمت‌دهی ستونی
        ws.cell(row=inflate_to_row, column=1)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


HEADER = ["تاریخ", "مبلغ", "مشتری"]
ROWS = [[f"2024-01-{(i % 28) + 1:02d}", 1000 + i, f"C{i:04d}"] for i in range(200)]


def test_bloated_dimension_reads_fast_and_exact():
    """dimension بادکرده (۲۰۰هزار ردیف اعلام‌شده، ۲۰۰ ردیف واقعی) → سریع و دقیق."""
    raw = _xlsx_bytes(ROWS, header=HEADER, inflate_to_row=200_000)
    conn = ExcelCsvConnector(content=raw, filename="bloated.xlsx")
    t0 = time.monotonic()
    result = conn.read(conn.list_sources()[0])
    elapsed = time.monotonic() - t0
    assert len(result.dataframe) == 200
    assert elapsed < 20, f"خواندن فایل بادکرده خیلی طول کشید: {elapsed:.1f}s"
    assert result.meta["warnings"] == []


def test_row_cap_with_persian_warning():
    raw = _xlsx_bytes(ROWS, header=HEADER)
    conn = ExcelCsvConnector(content=raw, filename="cap.xlsx")
    result = conn.read(None, max_rows=100)
    assert len(result.dataframe) == 100
    assert any("ردیف نخست" in w for w in result.meta["warnings"])


def test_small_gap_kept_large_gap_stops():
    """گپ کوتاه میانی مثل pd.read_excel به‌صورت NaN حفظ شود؛ گپ بزرگ = پایان داده."""
    from openpyxl import Workbook

    # گپ ۵ ردیفی وسط داده
    wb = Workbook()
    ws = wb.active
    ws.append(HEADER)
    for i in range(50):
        ws.append(["2024-01-01", 10 + i, f"A{i}"])
    for i in range(50):
        ws.cell(row=1 + 50 + 5 + i + 1, column=1, value="2024-02-01")
        ws.cell(row=1 + 50 + 5 + i + 1, column=2, value=20 + i)
        ws.cell(row=1 + 50 + 5 + i + 1, column=3, value=f"B{i}")
    buf = io.BytesIO()
    wb.save(buf)
    conn = ExcelCsvConnector(content=buf.getvalue(), filename="gap.xlsx")
    df = conn.read(None).dataframe
    control = pd.read_excel(io.BytesIO(buf.getvalue()), engine="openpyxl")
    assert len(df) == len(control) == 105  # ۵۰ + ۵ ردیف NaN + ۵۰

    # داده‌ی بعد از گپ بزرگ‌تر از EMPTY_RUN_LIMIT دور ریخته می‌شود
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(HEADER)
    for i in range(50):
        ws2.append(["2024-01-01", 10 + i, f"A{i}"])
    far = 1 + 50 + EMPTY_RUN_LIMIT + 50
    ws2.cell(row=far, column=1, value="ignored")
    buf2 = io.BytesIO()
    wb2.save(buf2)
    conn2 = ExcelCsvConnector(content=buf2.getvalue(), filename="farrow.xlsx")
    assert len(conn2.read(None).dataframe) == 50


def test_mapping_parity_old_vs_new_path():
    """گارد رگرسیون: نگاشت خودکار روی مسیر جدید عین مسیر pd.read_excel باشد."""
    df = generate_synthetic_sales().head(400)
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    raw = buf.getvalue()

    old = pd.read_excel(io.BytesIO(raw), engine="openpyxl")
    conn = ExcelCsvConnector(content=raw, filename="sales.xlsx")
    new = conn.read(None).dataframe

    assert list(old.columns) == list(new.columns)
    assert len(old) == len(new)

    mapper = SchemaMapper()
    s_old = mapper.auto_detect(old)
    s_new = mapper.auto_detect(new)
    assert s_old.mapping == s_new.mapping
    for role, g_old in s_old.guesses.items():
        g_new = s_new.guesses.get(role)
        if g_old.column is not None:
            assert g_new is not None
            assert abs(g_old.confidence - g_new.confidence) <= 0.01, role

    # dtypeهای کلیدی واقعاً برگشته باشند (نه object خام)
    rev_col = s_new.mapping[next(r for r in s_new.mapping if r.value == "REVENUE")]
    assert pd.api.types.is_numeric_dtype(new[rev_col])


def test_csv_path_unchanged():
    df = generate_synthetic_sales().head(300)
    raw = df.to_csv(index=False).encode("utf-8")
    conn = ExcelCsvConnector(content=raw, filename="sales.csv")
    result = conn.read(None)
    assert len(result.dataframe) == 300
    assert result.meta["warnings"] == []
    mapping = SchemaMapper().auto_detect(result.dataframe).mapping
    assert any(r.value == "REVENUE" for r in mapping)
    assert any(r.value == "DATE" for r in mapping)


def test_duplicate_and_unnamed_headers_mangled():
    res = StreamResult(rows=[[1, 2, 3, 4]], header=["مبلغ", None, "مبلغ", ""])
    df = finalize_frame(res)
    assert list(df.columns) == ["مبلغ", "Unnamed: 1", "مبلغ.1", "Unnamed: 3"]
    # دسترسی ستونی باید Series بدهد نه DataFrame (پیش‌نیاز mapper)
    assert isinstance(df["مبلغ"], pd.Series)


# --------------------------------------------------------- xlsb (iterator جعلی)
Cell = namedtuple("Cell", ["r", "c", "v"])


class _FakeSheet:
    def __init__(self, rows):
        self._rows = rows

    def rows(self, sparse=True):
        return iter(self._rows)


def test_xlsb_sparse_assembly_and_int_collapse():
    rows = [
        [Cell(0, 0, "تاریخ"), Cell(0, 2, "مبلغ")],          # گپ ستونی
        [Cell(1, 0, "2024-01-01"), Cell(1, 2, 7.0)],          # float صحیح → int
        [Cell(3, 0, "2024-01-02"), Cell(3, 2, 8.5)],          # گپ ردیفی (ردیف ۲ جا افتاده)
        # ردیف با عرض بادکرده: سلول‌های None تا ستون ۱۶۳۸۳ نباید عرض بسازند
        [Cell(4, 0, "2024-01-03"), Cell(4, 2, 9), Cell(4, 16383, None)],
    ]
    pairs = list(_xlsb_value_iter(_FakeSheet(rows)))
    assert pairs[0] == (0, ["تاریخ", None, "مبلغ"])
    assert pairs[1] == (1, ["2024-01-01", None, 7])          # int شد
    assert pairs[2] == (3, ["2024-01-02", None, 8.5])
    assert pairs[3][1] == ["2024-01-03", None, 9]            # عرض ۳ نه ۱۶۳۸۴

    dense = list(_densify(iter(pairs)))
    assert dense[2] == []                                     # ردیف جاافتاده → خالی


def test_xlsb_gap_counted_by_empty_run_guard():
    rows = [[Cell(0, 0, "H")], [Cell(1, 0, "a")]]
    # داده‌ی بعدی با گپ بزرگ‌تر از حد مجاز
    rows.append([Cell(1 + EMPTY_RUN_LIMIT + 5, 0, "b")])
    res = _collect(_densify(_xlsb_value_iter(_FakeSheet(rows))),
                   header_row=0, total=None, max_rows=10_000, progress=None)
    assert [r for r in res.rows if r] == [["a"]]


def test_collect_progress_called():
    calls: list[tuple[int, int | None]] = []
    values = iter([["h"]] + [["x"]] * 4500)
    _collect(values, header_row=0, total=9000, max_rows=10_000,
             progress=lambda done, total: calls.append((done, total)))
    assert calls, "progress هرگز صدا زده نشد"
    assert calls[-1][0] == 4500
    assert calls[-1][1] == 9000
