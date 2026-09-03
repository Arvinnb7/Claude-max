"""دروازه‌ی ساختاریِ ارسال: هیچ تخفیفی بدون تأییدِ انسان ارسال نمی‌شود.

تصمیمِ کاربر و §۲۰.۳. این فایل همان جمله را روی **مسیرِ واقعیِ ارسال** اثبات
می‌کند، نه روی عامل‌های موتور:

* قالبی که `{تخفیف}` دارد و آفرِ تأییدشده‌ای نیست ⇒ عضو ساختاراً ارسال
  نمی‌شود (نه «{تخفیف}» خام برای مشتری — رفتارِ قبلی — نه پیامِ بی‌تخفیف بی‌صدا).
* بعد از تأیید، همان عضو با پله‌ی تأییدشده می‌رود و `campaign_sends.offer_discount_bp`
  ثبت می‌شود (§۲۰.۲ «کدام آفر نشان داده شد»).
* گروه کنترل هرگز آفر نمی‌گیرد.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import select

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "src"))

from api.main import app  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

from mktcore.campaigns.assign import ARM_CONTROL, ARM_TREATMENT  # noqa: E402
from mktcore.db import session_scope  # noqa: E402
from mktcore.db.models import (  # noqa: E402
    CampaignMember,
    CampaignOpportunity,
    CampaignSend,
    Opportunity,
    OpportunityOffer,
)

from .conftest import poll_job, reset_contact_history  # noqa: E402
from .test_campaign_send import _enable_panel, _fake_panel  # noqa: E402

client = TestClient(app)


@pytest.fixture(autouse=True)
def _fresh_contact_history():
    """هر تست از وضعیتِ «کسی تازه تماس نگرفته» شروع می‌کند (خستگیِ تماسِ کمپین)."""
    reset_contact_history()
TEMPLATE = "سلام {نام}، این هفته {تخفیف} تخفیف برای شما داریم."


@pytest.fixture(scope="module")
def offers_ready() -> None:
    """تحلیلِ نمونه + نردبان + آستانه‌ای که مشتریانِ نمونه را «وابسته به تخفیف» کند."""
    r = client.post("/api/sample")
    data = r.json()
    mapping = {x["role"]: x["suggested"] for x in data["roles"] if x["suggested"]}
    r = client.post("/api/analyze", json={
        "session_id": data["session_id"], "mapping": mapping, "horizon": 3,
    })
    poll_job(client, r.json()["job_id"])

    assert client.put("/api/v1/margin-floor", json={"margin_floor_bp": 1_000}).status_code == 200
    # داده‌ی نمونه ~۸۰٪ خریدِ بدون تخفیف دارد؛ با این آستانه‌ها طبقه‌ی «پایین» می‌شود.
    r = client.put("/api/v1/offer-policy", json={
        "ladder_bp": [500, 1000, 1500],
        "full_price_high_bp": 9_900, "full_price_low_bp": 9_500, "full_price_min_lines": 2,
    })
    assert r.status_code == 200, r.text
    run = client.post("/api/v1/ops/jobs/opportunity_generation/run").json()
    assert run["status"] == "succeeded", run
    yield
    client.put("/api/v1/offer-policy", json={"ladder_bp": []})
    client.put("/api/v1/margin-floor", json={"margin_floor_bp": None})


def _campaign_with_offers() -> tuple[int, list[int], list[int]]:
    """کمپینی که دست‌کم یک عضوِ تیمارش پیشنهادِ تخفیف دارد."""
    r = client.post("/api/v1/campaigns", json={"name": "کمپین تخفیف", "holdout_pct": 20, "limit": 60})
    assert r.status_code == 200, r.text
    campaign_id = r.json()["id"]
    with session_scope() as session:
        arms = session.execute(
            select(CampaignMember.customer_id, CampaignMember.arm)
            .where(CampaignMember.campaign_id == campaign_id)
        ).all()
        treatment = [int(c) for c, a in arms if a == ARM_TREATMENT]
        suggested = session.scalars(
            select(CampaignOpportunity.opportunity_id)
            .join(OpportunityOffer, OpportunityOffer.opportunity_id == CampaignOpportunity.opportunity_id)
            .where(
                CampaignOpportunity.campaign_id == campaign_id,
                CampaignOpportunity.customer_id.in_(treatment),
                OpportunityOffer.status == OpportunityOffer.STATUS_SUGGESTED,
            )
        ).all()
    assert suggested, "هیچ عضوِ تیماری پیشنهادِ تخفیف ندارد؛ فیکسچر معنا ندارد"
    return campaign_id, treatment, [int(x) for x in suggested]


def _sends(campaign_id: int) -> list[CampaignSend]:
    with session_scope() as session:
        rows = session.scalars(
            select(CampaignSend).where(CampaignSend.campaign_id == campaign_id)
        ).all()
        session.expunge_all()
    return rows


def test_unapproved_offers_are_structurally_unsendable(offers_ready, monkeypatch):
    campaign_id, _treatment, _suggested = _campaign_with_offers()
    sink: list[str] = []
    _fake_panel(monkeypatch, sink)
    _enable_panel(monkeypatch)

    r = client.post(f"/api/v1/campaigns/{campaign_id}/send", json={
        "template": TEMPLATE, "dry_run": False, "confirm": True,
    })
    assert r.status_code == 200, r.text
    send = r.json()["send"]

    assert send["ارسال‌شده"] == 0
    assert send["بدون_تأیید_تخفیف"] >= 1
    assert not sink, "هیچ پیامی نباید به پنل رفته باشد"
    rows = _sends(campaign_id)
    assert rows and all(row.status == CampaignSend.STATUS_SKIPPED for row in rows)
    assert all("تأییدشده" in (row.status_detail_fa or "") for row in rows)
    assert all("{تخفیف}" not in (row.message_text or "") for row in rows)
    with session_scope() as session:
        stamped = session.scalar(
            select(CampaignMember.id).where(
                CampaignMember.campaign_id == campaign_id,
                CampaignMember.exposure_at.isnot(None),
            )
        )
    assert stamped is None, "عضوِ ارسال‌نشده نباید مهرِ تماس بخورد"


def test_an_approved_offer_is_sent_with_its_rung_and_logged(offers_ready, monkeypatch):
    campaign_id, _treatment, suggested = _campaign_with_offers()
    sink: list[str] = []
    _fake_panel(monkeypatch, sink)
    _enable_panel(monkeypatch)

    approve = client.post(
        f"/api/v1/opportunities/{suggested[0]}/offer/approve",
        json={"decided_by": "مدیر فروش"},
    )
    assert approve.status_code == 200, approve.text
    assert approve.json()["sendable"] is True

    r = client.post(f"/api/v1/campaigns/{campaign_id}/send", json={
        "template": TEMPLATE, "dry_run": False, "confirm": True,
    })
    send = r.json()["send"]
    assert send["ارسال‌شده"] == 1, send
    assert len(sink) == 1

    sent = [row for row in _sends(campaign_id) if row.status == CampaignSend.STATUS_SENT]
    assert len(sent) == 1
    assert sent[0].offer_discount_bp == 500
    assert "5٪" in sent[0].message_text and "{تخفیف}" not in sent[0].message_text
    with session_scope() as session:
        member = session.scalar(
            select(CampaignMember).where(
                CampaignMember.campaign_id == campaign_id,
                CampaignMember.customer_id == sent[0].customer_id,
            )
        )
        assert member.offer_discount_bp == 500
        assert member.exposure_at is not None


def test_control_arm_never_carries_an_offer(offers_ready):
    campaign_id, _treatment, _suggested = _campaign_with_offers()
    with session_scope() as session:
        control_offers = session.scalar(
            select(CampaignMember.id).where(
                CampaignMember.campaign_id == campaign_id,
                CampaignMember.arm == ARM_CONTROL,
                CampaignMember.offer_discount_bp.isnot(None),
            )
        )
        control_sends = session.scalar(
            select(CampaignSend.id)
            .join(CampaignMember, CampaignMember.customer_id == CampaignSend.customer_id)
            .where(
                CampaignSend.campaign_id == campaign_id,
                CampaignMember.campaign_id == campaign_id,
                CampaignMember.arm == ARM_CONTROL,
            )
        )
    assert control_offers is None
    assert control_sends is None


def test_export_shows_the_approved_discount_and_never_a_raw_placeholder(offers_ready):
    campaign_id, _treatment, suggested = _campaign_with_offers()
    client.post(f"/api/v1/opportunities/{suggested[0]}/offer/approve", json={})

    r = client.get(f"/api/v1/campaigns/{campaign_id}/export")
    assert r.status_code == 200
    sheet = openpyxl.load_workbook(io.BytesIO(r.content))["فهرست تماس"]
    headers = [c.value for c in sheet[1]]
    assert "تخفیف تأییدشده" in headers
    col = headers.index("تخفیف تأییدشده") + 1
    values = [sheet.cell(row=i, column=col).value for i in range(2, sheet.max_row + 1)]
    assert "5٪" in values
    assert "—" in values
    for row in sheet.iter_rows(min_row=2, values_only=True):
        assert all("{تخفیف}" not in str(cell) for cell in row)


# ═══════════════════════════════════ یافته‌های بازبینی خصمانه
def test_a_template_without_the_placeholder_does_not_stamp_an_offer(offers_ready, monkeypatch):
    """آفرِ تأییدشده ولی قالبِ بی‌`{تخفیف}` ⇒ مشتری تخفیفی ندیده ⇒ هیچ مهرِ آفری."""
    campaign_id, _treatment, suggested = _campaign_with_offers()
    sink: list[str] = []
    _fake_panel(monkeypatch, sink)
    _enable_panel(monkeypatch)
    assert client.post(f"/api/v1/opportunities/{suggested[0]}/offer/approve", json={}).status_code == 200

    r = client.post(f"/api/v1/campaigns/{campaign_id}/send", json={
        "template": "سلام {نام}، کالای تازه رسید.", "dry_run": False, "confirm": True,
    })
    send = r.json()["send"]
    assert send["ارسال‌شده"] >= 1 and send["بدون_تأیید_تخفیف"] == 0

    sent = [row for row in _sends(campaign_id) if row.status == CampaignSend.STATUS_SENT]
    assert sent and all(row.offer_discount_bp is None for row in sent)
    with session_scope() as session:
        stamped = session.scalar(
            select(CampaignMember.id).where(
                CampaignMember.campaign_id == campaign_id,
                CampaignMember.offer_discount_bp.isnot(None),
            )
        )
    assert stamped is None


def test_a_member_exposed_by_export_first_still_records_the_sent_rung(offers_ready, monkeypatch):
    """اول دانلودِ فهرست (بدون آفر)، بعد تأیید، بعد پیامک با تخفیف ⇒ پله‌ی واقعاً ارسال‌شده ثبت شود."""
    campaign_id, _treatment, suggested = _campaign_with_offers()
    assert client.get(f"/api/v1/campaigns/{campaign_id}/export").status_code == 200
    assert client.post(f"/api/v1/opportunities/{suggested[0]}/offer/approve", json={}).status_code == 200

    sink: list[str] = []
    _fake_panel(monkeypatch, sink)
    _enable_panel(monkeypatch)
    r = client.post(f"/api/v1/campaigns/{campaign_id}/send", json={
        "template": TEMPLATE, "dry_run": False, "confirm": True,
    })
    assert r.json()["send"]["ارسال‌شده"] >= 1

    # تأییدهای تست‌های قبلی روی همین دفتر کل می‌مانند؛ پس فقط عضوی را می‌سنجیم
    # که همین تست تأیید کرد.
    with session_scope() as session:
        customer_id = session.scalar(
            select(Opportunity.customer_id).where(Opportunity.id == suggested[0])
        )
        sent = session.scalar(
            select(CampaignSend).where(
                CampaignSend.campaign_id == campaign_id,
                CampaignSend.customer_id == customer_id,
                CampaignSend.status == CampaignSend.STATUS_SENT,
            )
        )
        assert sent is not None and sent.offer_discount_bp == 500
        member = session.scalar(
            select(CampaignMember).where(
                CampaignMember.campaign_id == campaign_id,
                CampaignMember.customer_id == customer_id,
            )
        )
        assert member.exposure_channel == "excel_export", "مهرِ اولین تماس یک‌بارمصرف می‌ماند"
        assert member.offer_discount_bp == 500, "ولی پله‌ی ارسال‌شده باید ثبت شود"
