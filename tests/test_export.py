"""تست‌های خروجی اکسل بخش‌های داشبورد."""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "src"))

from api.main import app  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

from mktcore.analysis.segmentation import RFM_SEGMENT_NAMES  # noqa: E402

from .conftest import poll_job  # noqa: E402

client = TestClient(app)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _analyzed_session() -> str:
    r = client.post("/api/sample")
    data = r.json()
    roles = {x["role"]: x["suggested"] for x in data["roles"]}
    mapping = {role: col for role, col in roles.items() if col}
    r = client.post("/api/analyze", json={
        "session_id": data["session_id"], "mapping": mapping, "horizon": 4,
    })
    poll_job(client, r.json()["job_id"])
    return data["session_id"]


def _fetch(sid: str, section: str):
    r = client.get(f"/api/export?session_id={sid}&section={section}")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(_XLSX_MIME)
    assert "filename*=UTF-8''" in r.headers["content-disposition"]
    return openpyxl.load_workbook(io.BytesIO(r.content))


def test_export_segments_with_phones():
    sid = _analyzed_session()
    wb = _fetch(sid, "segments")
    valid_names = set(RFM_SEGMENT_NAMES.values())
    assert set(wb.sheetnames) <= valid_names
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    assert "کد مشتری" in headers and "موبایل" in headers
    phone_col = headers.index("موبایل") + 1
    phones = [ws.cell(row=i, column=phone_col).value for i in range(2, min(ws.max_row, 30) + 1)]
    assert any(p for p in phones), "هیچ شماره‌ی موبایلی در خروجی سگمنت نیست"


def test_export_next_purchase_probabilities():
    sid = _analyzed_session()
    wb = _fetch(sid, "next_purchase")
    ws = wb["پیش‌بینی خرید بعدی"]
    headers = [c.value for c in ws[1]]
    assert "احتمال خرید ۳۰ روز" in headers and "موبایل" in headers
    p_col = headers.index("احتمال خرید ۳۰ روز") + 1
    vals = [ws.cell(row=i, column=p_col).value for i in range(2, ws.max_row + 1)]
    nums = [v for v in vals if isinstance(v, (int, float))]
    assert nums, "هیچ احتمال عددی در خروجی نیست"
    assert all(0.0 <= v <= 1.0 for v in nums)


def test_export_products_full_list():
    sid = _analyzed_session()
    wb = _fetch(sid, "products")
    ws = wb["محصولات پرفروش"]
    headers = [c.value for c in ws[1]]
    assert "محصول" in headers and "کلاس ABC" in headers
    assert ws.max_row >= 2


def test_export_diagnostics_multi_sheet():
    sid = _analyzed_session()
    wb = _fetch(sid, "diagnostics")
    assert "پیشنهاد تأمین" in wb.sheetnames
    assert any(n.startswith("خلاصه") or n == "برنامه هفته" for n in wb.sheetnames)


def test_export_invalid_section_and_session():
    sid = _analyzed_session()
    assert client.get(f"/api/export?session_id={sid}&section=nope").status_code == 400
    assert client.get("/api/export?session_id=missing&section=products").status_code == 404
