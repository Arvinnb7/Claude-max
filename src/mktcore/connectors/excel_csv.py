"""کانکتور کامل اکسل/CSV: تشخیص encoding و جداکننده، چندشیت، هدر چندردیفه."""

from __future__ import annotations

import csv
import io
from collections.abc import Callable
from pathlib import Path

import pandas as pd

from ..config import get_settings
from .base import ConnectorResult, DataConnector
from .excel_stream import (
    _cap_warning,
    finalize_frame,
    read_xlsb_stream,
    read_xlsx_stream,
)

_EXCEL_EXT = {".xlsx", ".xlsm", ".xls", ".xlsb"}
_CSV_EXT = {".csv", ".txt", ".tsv"}

# فایل‌های CSV کوچک‌تر از این حجم با یک read_csv تکی خوانده می‌شوند (رفتار قبلی)
_CSV_SINGLE_READ_BYTES = 5 * 1024 * 1024


def _excel_engine(ext: str) -> str:
    """انتخاب موتور مناسب خواندن اکسل بر اساس پسوند."""
    if ext == ".xls":
        return "xlrd"
    if ext == ".xlsb":
        return "pyxlsb"  # اکسل باینری (نیازمند بسته‌ی pyxlsb)
    return "openpyxl"


def _sniff_csv(raw: bytes) -> tuple[str, str]:
    """تشخیص encoding و جداکننده‌ی یک فایل متنی."""
    encoding = "utf-8"
    for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
        try:
            raw.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    text_sample = raw[:8192].decode(encoding, errors="ignore")
    try:
        dialect = csv.Sniffer().sniff(text_sample, delimiters=",;\t|")
        sep = dialect.delimiter
    except csv.Error:
        # شمارش ساده‌ی نامزدها
        counts = {d: text_sample.count(d) for d in [",", ";", "\t", "|"]}
        sep = max(counts, key=counts.get) if any(counts.values()) else ","
    return encoding, sep


class ExcelCsvConnector(DataConnector):
    """خواندن فایل‌های اکسل و CSV از مسیر یا بایت‌های آپلودشده."""

    name = "excel_csv"

    def __init__(self, *, path: str | Path | None = None, content: bytes | None = None,
                 filename: str | None = None) -> None:
        if path is None and content is None:
            raise ValueError("باید path یا content داده شود.")
        self.path = Path(path) if path else None
        self.content = content
        self.filename = filename or (self.path.name if self.path else "uploaded")
        self._ext = self._detect_format(Path(self.filename).suffix.lower())

    def _detect_format(self, ext: str) -> str:
        """قالب واقعی فایل از **امضای بایت‌ها**، با بازگشت به پسوند.

        پسوند دروغ می‌گوید: کاربران فایل `.xls` را با نام `.xlsx` ذخیره می‌کنند
        (اکسل هشدار می‌دهد و آن‌ها رد می‌کنند)، سامانه‌ها فایل بدون پسوند
        می‌فرستند، و مرورگر گاهی `.csv` را `.xls` نام می‌گذارد. تشخیص از محتوا
        یعنی فایل درست خوانده می‌شود یا با خطای روشن رد می‌شود — نه اینکه با
        parser اشتباه به هم بریزد.
        """
        try:
            head = self._peek(8)
        except OSError:
            return ext

        if head[:4] == b"PK\x03\x04":
            # ZIP: هم xlsx/xlsm است هم xlsb — تفاوت در محتوای داخل آرشیو
            if ext in (".xlsx", ".xlsm", ".xlsb"):
                return ext
            return self._zip_flavor()
        if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            return ".xls"  # فرمت قدیمی OLE2
        if ext in _EXCEL_EXT:
            # پسوند اکسل ولی امضای غیراکسل → احتمالاً CSV با نام اشتباه
            return ".csv"
        return ext or ".csv"

    def _zip_flavor(self) -> str:
        """داخل آرشیو را نگاه می‌کند تا xlsb را از xlsx جدا کند."""
        import zipfile

        try:
            with zipfile.ZipFile(io.BytesIO(self._bytes())) as zf:
                names = set(zf.namelist())
        except (zipfile.BadZipFile, OSError):
            return ".xlsx"
        if any(n.endswith(".bin") and "worksheets/" in n for n in names):
            return ".xlsb"
        return ".xlsx"

    def _peek(self, n: int) -> bytes:
        """چند بایت اول، بدون خواندن کل فایل از دیسک."""
        if self.content is not None:
            return self.content[:n]
        if self.path is None:  # pragma: no cover - سازنده جلوی این را می‌گیرد
            return b""
        with self.path.open("rb") as fh:
            return fh.read(n)

    def _bytes(self) -> bytes:
        if self.content is not None:
            return self.content
        if self.path is None:  # pragma: no cover - سازنده جلوی این را می‌گیرد
            raise ValueError("نه مسیر فایل و نه محتوای آن در دسترس است.")
        return self.path.read_bytes()

    def list_sources(self) -> list[str]:
        """نام شیت‌ها (اکسل) یا نام فایل (CSV) — فقط متادیتا، بدون پارس کامل شیت."""
        if self._ext in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(self._bytes()), read_only=True)
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()
        if self._ext == ".xlsb":
            from pyxlsb import open_workbook

            with open_workbook(io.BytesIO(self._bytes())) as wb:
                return list(wb.sheets)
        if self._ext == ".xls":
            import xlrd

            book = xlrd.open_workbook(file_contents=self._bytes(), on_demand=True)
            return list(book.sheet_names())
        if self._ext in _CSV_EXT:
            return [self.filename]
        raise ValueError(f"پسوند پشتیبانی‌نشده: {self._ext}")

    def read(self, source: str | None = None, *, header_row: int = 0,
             progress: Callable[[int, int | None], None] | None = None,
             max_rows: int | None = None, **kwargs) -> ConnectorResult:
        """خواندن یک شیت/فایل و بازگرداندن DataFrame خام.

        Args:
            source: نام شیت (اکسل) — در نبود، اولین شیت.
            header_row: شماره‌ی ردیف هدر (۰-مبنا) برای فایل‌های با هدر چندردیفه.
            progress: callback اختیاری (ردیف‌های خوانده‌شده، کل اعلام‌شده یا None).
            max_rows: سقف ردیف؛ پیش‌فرض از تنظیمات (MKT_MAX_ROWS).
        """
        raw = self._bytes()
        cap = max_rows or get_settings().mkt_max_rows
        warnings: list[str] = []

        if self._ext in (".xlsx", ".xlsm"):
            res = read_xlsx_stream(raw, source, header_row=header_row,
                                   max_rows=cap, progress=progress)
            df = finalize_frame(res)
            warnings = res.warnings
            meta = {"format": "excel", "sheet": source}
        elif self._ext == ".xlsb":
            res = read_xlsb_stream(raw, source, header_row=header_row,
                                   max_rows=cap, progress=progress)
            df = finalize_frame(res)
            warnings = res.warnings
            meta = {"format": "excel", "sheet": source}
        elif self._ext == ".xls":
            # xlrd کل فایل را هنگام باز کردن پارس می‌کند و سقف فرمت ۶۵۵۳۶ ردیف است؛
            # استریم معنایی ندارد — فقط سقف و progress ابتدا/انتها.
            if progress:
                progress(0, None)
            df = pd.read_excel(
                io.BytesIO(raw),
                sheet_name=source if source else 0,
                header=header_row,
                engine="xlrd",
            )
            if len(df) > cap:
                df = df.head(cap)
                warnings.append(_cap_warning(cap))
            if progress:
                progress(len(df), len(df))
            meta = {"format": "excel", "sheet": source}
        elif self._ext in _CSV_EXT:
            encoding, sep = _sniff_csv(raw)
            if len(raw) <= _CSV_SINGLE_READ_BYTES:
                df = pd.read_csv(io.BytesIO(raw), encoding=encoding, sep=sep,
                                 header=header_row)
                if len(df) > cap:
                    df = df.head(cap)
                    warnings.append(_cap_warning(cap))
            else:
                parts: list[pd.DataFrame] = []
                n = 0
                reader = pd.read_csv(io.BytesIO(raw), encoding=encoding, sep=sep,
                                     header=header_row, chunksize=50_000)
                for chunk in reader:
                    take = min(len(chunk), cap - n)
                    parts.append(chunk.head(take))
                    n += take
                    if progress:
                        progress(n, None)
                    if n >= cap:
                        warnings.append(_cap_warning(cap))
                        break
                df = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
            meta = {"format": "csv", "encoding": encoding, "sep": sep}
        else:
            raise ValueError(f"پسوند پشتیبانی‌نشده: {self._ext}")

        # حذف ستون‌های کاملاً بی‌نام/خالی
        df = df.dropna(axis=1, how="all")
        df.columns = [str(c).strip() for c in df.columns]
        meta["warnings"] = warnings
        return ConnectorResult(dataframe=df, source_name=source or self.filename, meta=meta)
