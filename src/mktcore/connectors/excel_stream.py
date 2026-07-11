"""خواندن استریمی اکسل با گزارش پیشرفت و گاردهای محدوده‌ی بادکرده.

فایل‌های واقعی فروش خیلی وقت‌ها used-range بادکرده دارند (فرمت‌دهی ستونی →
dimension چند صدهزار ردیفِ خالی). pd.read_excel کل آن محدوده را می‌خواند که
هم بسیار کند است و هم حافظه را منفجر می‌کند. این ماژول ردیف‌به‌ردیف می‌خواند،
با دو گارد: توقف بعد از EMPTY_RUN_LIMIT ردیف خالی متوالی، و سقف max_rows.
"""

from __future__ import annotations

import datetime as _dt
import io
from collections.abc import Callable, Iterator
from dataclasses import dataclass, field

import pandas as pd

from ..locale_fa import format_number_fa

# (تعداد ردیف خوانده‌شده، کل ردیف‌های اعلام‌شده یا None)
ReadProgress = Callable[[int, int | None], None]

# بعد از این تعداد ردیف کاملاً خالیِ متوالی، داده تمام‌شده تلقی می‌شود.
# گپ‌های واقعی (ردیف‌های جداکننده) ۱ تا چند ده ردیف‌اند؛ ۳۰۰ ردیف خالی پشت‌سرهم
# فقط در فایل‌های با محدوده‌ی بادکرده دیده می‌شود.
EMPTY_RUN_LIMIT = 300

# هر چند ردیف یک‌بار progress گزارش شود
_PROGRESS_EVERY = 2000


@dataclass
class StreamResult:
    """خروجی خام خواندن استریمی — قبل از ساخت DataFrame."""

    rows: list[list]
    header: list
    warnings: list[str] = field(default_factory=list)
    truncated: bool = False


def _is_empty_row(values) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def _cap_warning(max_rows: int) -> str:
    n = format_number_fa(max_rows)
    return (
        f"فایل بیش از {n} ردیف دارد؛ فقط {n} ردیف نخست خوانده شد و "
        "ادامه‌ی فایل نادیده گرفته شد."
    )


def _collect(value_iter: Iterator, *, header_row: int, total: int | None,
             max_rows: int, progress: ReadProgress | None) -> StreamResult:
    """حلقه‌ی مشترک جمع‌آوری ردیف‌ها با گاردهای empty-run و سقف ردیف.

    گپ‌های خالی کوتاه (کمتر از EMPTY_RUN_LIMIT) مثل pd.read_excel به‌صورت
    ردیف NaN حفظ می‌شوند؛ ردیف‌های خالی انتهایی دور ریخته می‌شوند.
    """
    header: list | None = None
    rows: list[list] = []
    pending_empty = 0
    warnings: list[str] = []
    truncated = False

    for i, values in enumerate(value_iter):
        if i < header_row:
            continue
        if i == header_row:
            header = list(values)
            continue
        if _is_empty_row(values):
            pending_empty += 1
            if pending_empty >= EMPTY_RUN_LIMIT:
                break
            continue
        if pending_empty:
            rows.extend([[] for _ in range(pending_empty)])
            pending_empty = 0
        rows.append(list(values))
        if len(rows) >= max_rows:
            truncated = True
            warnings.append(_cap_warning(max_rows))
            break
        if progress and len(rows) % _PROGRESS_EVERY == 0:
            progress(len(rows), total)

    if progress:
        progress(len(rows), total)
    return StreamResult(rows=rows, header=header or [], warnings=warnings,
                        truncated=truncated)


# --------------------------------------------------------------- xlsx/xlsm
def read_xlsx_stream(raw: bytes, sheet: str | None, *, header_row: int,
                     max_rows: int, progress: ReadProgress | None) -> StreamResult:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True,
                       keep_links=False)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
        total = ws.max_row  # بُعد اعلام‌شده — ممکن است بادکرده باشد
        return _collect(ws.iter_rows(values_only=True), header_row=header_row,
                        total=total, max_rows=max_rows, progress=progress)
    finally:
        wb.close()  # حالت read_only فایل را باز نگه می‌دارد؛ بستن الزامی است


# -------------------------------------------------------------------- xlsb
def _xlsb_value_iter(ws) -> Iterator[tuple[int, list]]:
    """(شماره‌ی ردیف صفر-مبنا، مقادیر) از یک شیت pyxlsb.

    چینش بر اساس cell.c تا به ترتیب/تراکم خروجی pyxlsb تکیه نکنیم؛ عرض هر
    ردیف = آخرین ستون غیرخالی + ۱ (هرس عرض بادکرده‌ی dimension تا 16384).
    مثل pandas، floatهای صحیح به int جمع می‌شوند.
    """
    for row in ws.rows(sparse=True):
        if not row:
            continue
        r = row[0].r
        vals: list = []
        for cell in row:
            v = cell.v
            if v is None:
                continue
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            if cell.c >= len(vals):
                vals.extend([None] * (cell.c - len(vals)))
                vals.append(v)
            else:
                vals[cell.c] = v
        yield r, vals


def _densify(indexed_iter: Iterator[tuple[int, list]]) -> Iterator[list]:
    """گپ شماره‌ردیف‌ها را به ردیف خالی تبدیل می‌کند تا در گارد empty-run شمرده شوند."""
    prev = -1
    for r, vals in indexed_iter:
        for _ in range(r - prev - 1):
            yield []
        yield vals
        prev = r


def read_xlsb_stream(raw: bytes, sheet: str | None, *, header_row: int,
                     max_rows: int, progress: ReadProgress | None) -> StreamResult:
    from pyxlsb import open_workbook

    with open_workbook(io.BytesIO(raw)) as wb:
        name_or_idx = sheet if sheet else 1  # pyxlsb یک-مبنا است
        with wb.get_sheet(name_or_idx) as ws:
            total = None
            dim = getattr(ws, "dimension", None)
            if dim is not None and getattr(dim, "h", 0):
                total = int(dim.h)
            return _collect(_densify(_xlsb_value_iter(ws)), header_row=header_row,
                            total=total, max_rows=max_rows, progress=progress)


# ------------------------------------------------------- ساخت DataFrame نهایی
def _mangle_headers(cells: list) -> list[str]:
    """نام‌گذاری ستون‌ها آینه‌ی pd.read_excel: بی‌نام‌ها Unnamed: i و تکراری‌ها پسوند .n"""
    names: list[str] = []
    seen: dict[str, int] = {}
    for i, c in enumerate(cells):
        name = "" if c is None else str(c).strip()
        if not name or name.lower() in ("nan", "none"):
            name = f"Unnamed: {i}"
        n = seen.get(name, 0)
        seen[name] = n + 1
        names.append(name if n == 0 else f"{name}.{n}")
    return names


def _infer_like_read_excel(df: pd.DataFrame) -> pd.DataFrame:
    """بازگرداندن dtypeهای عددی/تاریخ که pd.read_excel به‌طور طبیعی می‌ساخت.

    ستون‌های شامل مقادیر تایپ‌شده‌ی عددی → float/int واقعی (حافظه و parquet)؛
    تاریخ‌های سلولی → datetime64. بقیه با infer_objects (سازگار pandas 2 و 3).
    """
    out: dict[str, pd.Series] = {}
    for col in df.columns:
        s = df[col]
        nn = s.dropna()
        if len(nn) and nn.map(
            lambda v: isinstance(v, (int, float)) and not isinstance(v, bool)
        ).all():
            out[col] = pd.to_numeric(s, errors="coerce")
        elif len(nn) and nn.map(
            lambda v: isinstance(v, (_dt.datetime, _dt.date))
        ).all():
            out[col] = pd.to_datetime(s, errors="coerce")
        else:
            out[col] = s.infer_objects()
    return pd.DataFrame(out, index=df.index)


def finalize_frame(res: StreamResult) -> pd.DataFrame:
    """StreamResult → DataFrame با هدر، عرض یکنواخت و dtypeهای استنتاج‌شده."""
    names = _mangle_headers(res.header)
    width = len(names)
    rows = [
        r + [None] * (width - len(r)) if len(r) < width else r[:width]
        for r in res.rows
    ]
    df = pd.DataFrame(rows, columns=names, dtype=object)
    return _infer_like_read_excel(df)
